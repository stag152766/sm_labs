import numpy as np
from openpyxl import load_workbook

table = []
row = []


def expo_random(time):
    average = 0
    n = 100

    for i in range(n):
        r = np.random.random()
        t = -time * np.log(r)
        average += (t / n)
    return round(average, 2)


def fill_row():
    row.clear()
    for i in range(5):
        r = expo_random(5)
        row.append(r)
    table.append(row)


# fill_row()
# print(table)
#
# fill_row()
# print(table)


# f = open('TT2.txt', 'r', encoding='utf-8')
#
# for line in f:
#     print(line)


# wb = load_workbook('TT2.xlsx')
# print(wb.get_sheet_names)
#
# sheet = wb.get_sheet_by_name('TT2')
#
# for i in range(1,100):
#     Ti1 = sheet.cell(i,1).value
#     print(Ti1)

a = np.zeros((6, 100), int)

a[1,1] = 1
print(a)