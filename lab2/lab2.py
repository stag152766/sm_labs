import numpy as np
from openpyxl import load_workbook

wb = load_workbook('TT2.xlsx')
sheet = wb.get_sheet_by_name('TT')

n = 100 + 1
m = 6

N = np.zeros((m, n), int)
T = np.zeros((m, n), int)
total_time = [0]
difference = [0]
T[0, 0] = 0

for i in range(1, n):
    N[0, i] = T[0, i - 1]
    U = sheet.cell(i, 1).value
    T[0, i] = N[0, i] + int(U)

for j in range(1, m):
    N[j, 1] = T[j - 1, 1]
    U = sheet.cell(1, j + 1).value
    T[j, 1] = N[j, 1] + int(U)

for i in range(2, n):
    for j in range(1, m):
        N[j, i] = max(T[j, i - 1], T[j - 1, i])
        U = sheet.cell(i, j + 1).value
        T[j, i] = N[j, i] + int(U)

print("N")

for i in range(len(N)):
    current_row = N[i]
    for j in current_row:
        print("%7s" % (j), end="\t\t")
    print('\r')

print("T")

for i in range(1, n):
    U = sheet.cell(i, 7).value
    total_time.append(int(U))

for i in range(1, n):
    delta = T[5, i] - total_time[i]
    difference.append(delta)

max_total_time = max(difference)
sum_total_time = sum(difference)

for i in range(len(T)):
    current_row = T[i]
    for j in current_row:
        print("%7s" % (j), end="\t\t")
    print('\r')
print("Суммарное время обработки:", T[-1][-1]);
print("Суммарное опоздание деталей:", sum_total_time)
print("Максимальное опоздание: ", max_total_time)
